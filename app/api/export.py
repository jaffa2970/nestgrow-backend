import io
import logging
from datetime import datetime, timedelta, timezone

import openpyxl
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.auth import require_admin
from app.database import get_db
from app.models import Culla, Irrigazione, Lettura, Zona

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/export", tags=["export"])

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2d6a4f")
_ALT_FILL = PatternFill("solid", fgColor="F5F5F5")
_CENTER = Alignment(horizontal="center")


def _style_sheet(ws, col_widths: list[float]) -> None:
    ws.freeze_panes = "A2"
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALT_FILL and Alignment(horizontal="left")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill


def _cutoff(giorni: int) -> datetime | None:
    if giorni <= 0:
        return None
    return datetime.now(timezone.utc) - timedelta(days=giorni)


def _filename(giorni: int) -> str:
    return f"nestgrow_export_{datetime.now().strftime('%Y%m%d')}.xlsx"


def _stream(wb: openpyxl.Workbook, giorni: int) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={_filename(giorni)}"},
    )


async def _build_letture_sheet(ws, db: AsyncSession, giorni: int) -> None:
    ws.title = "Letture umidità"
    ws.append(["Data/Ora", "Culla", "Zona", "Coltura", "Umidità %", "Livello Serbatoio %"])

    q = (
        select(Lettura.ts, Culla.nome, Zona.numero_zona, Zona.descrizione_coltura,
               Lettura.umidita_pct, Lettura.livello_serbatoio_pct)
        .join(Zona, Zona.id == Lettura.zona_id)
        .join(Culla, Culla.id == Zona.culla_id)
        .order_by(Lettura.ts.desc())
    )
    cutoff = _cutoff(giorni)
    if cutoff:
        q = q.where(Lettura.ts >= cutoff)

    rows = (await db.execute(q)).all()
    for r in rows:
        ws.append([
            r.ts.strftime("%d/%m/%Y %H:%M:%S") if r.ts else "",
            r.nome,
            f"Zona {r.numero_zona}",
            r.descrizione_coltura or "",
            round(r.umidita_pct, 1) if r.umidita_pct is not None else "",
            round(r.livello_serbatoio_pct, 1) if r.livello_serbatoio_pct is not None else "",
        ])

    _style_sheet(ws, [22, 20, 10, 24, 13, 20])


async def _build_irrigazioni_sheet(ws, db: AsyncSession, giorni: int) -> None:
    ws.title = "Irrigazioni"
    ws.append([
        "Data/Ora inizio", "Data/Ora fine", "Culla", "Zona", "Coltura",
        "Durata (sec)", "Umidità pre %", "Umidità post %", "Trigger", "Esito",
    ])

    q = (
        select(Irrigazione.ts_inizio, Irrigazione.ts_fine, Culla.nome,
               Zona.numero_zona, Zona.descrizione_coltura,
               Irrigazione.durata_sec, Irrigazione.umidita_pre,
               Irrigazione.umidita_post, Irrigazione.trigger, Irrigazione.esito)
        .join(Zona, Zona.id == Irrigazione.zona_id)
        .join(Culla, Culla.id == Zona.culla_id)
        .order_by(Irrigazione.ts_inizio.desc())
    )
    cutoff = _cutoff(giorni)
    if cutoff:
        q = q.where(Irrigazione.ts_inizio >= cutoff)

    rows = (await db.execute(q)).all()
    for r in rows:
        ws.append([
            r.ts_inizio.strftime("%d/%m/%Y %H:%M:%S") if r.ts_inizio else "",
            r.ts_fine.strftime("%d/%m/%Y %H:%M:%S") if r.ts_fine else "",
            r.nome,
            f"Zona {r.numero_zona}",
            r.descrizione_coltura or "",
            r.durata_sec or "",
            round(r.umidita_pre, 1) if r.umidita_pre is not None else "",
            round(r.umidita_post, 1) if r.umidita_post is not None else "",
            r.trigger or "",
            r.esito or "",
        ])

    _style_sheet(ws, [22, 22, 20, 10, 24, 13, 13, 14, 12, 12])


async def _build_riepilogo_sheet(ws, db: AsyncSession, giorni: int) -> None:
    ws.title = "Riepilogo per zona"
    ws.append([
        "Culla", "Zona", "Coltura",
        "N. irrigazioni", "Durata media (sec)",
        "Umidità media %", "Umidità min %", "Umidità max %",
    ])

    cutoff = _cutoff(giorni)

    # Aggregated letture stats per zona
    letture_q = select(
        Lettura.zona_id,
        func.avg(Lettura.umidita_pct).label("media"),
        func.min(Lettura.umidita_pct).label("minimo"),
        func.max(Lettura.umidita_pct).label("massimo"),
    ).where(Lettura.umidita_pct.is_not(None)).group_by(Lettura.zona_id)
    if cutoff:
        letture_q = letture_q.where(Lettura.ts >= cutoff)
    letture_stats = {r.zona_id: r for r in (await db.execute(letture_q)).all()}

    # Aggregated irrigazioni stats per zona
    irr_q = select(
        Irrigazione.zona_id,
        func.count(Irrigazione.id).label("n"),
        func.avg(Irrigazione.durata_sec).label("durata_media"),
    ).group_by(Irrigazione.zona_id)
    if cutoff:
        irr_q = irr_q.where(Irrigazione.ts_inizio >= cutoff)
    irr_stats = {r.zona_id: r for r in (await db.execute(irr_q)).all()}

    # All active zone with culla name
    zone_q = (
        select(Zona, Culla.nome.label("culla_nome"))
        .join(Culla, Culla.id == Zona.culla_id)
        .where(Culla.attiva == True)
        .order_by(Culla.nome, Zona.numero_zona)
    )
    zone = (await db.execute(zone_q)).all()

    for zona, culla_nome in zone:
        ls = letture_stats.get(zona.id)
        ir = irr_stats.get(zona.id)
        ws.append([
            culla_nome,
            f"Zona {zona.numero_zona}",
            zona.descrizione_coltura or "",
            ir.n if ir else 0,
            round(float(ir.durata_media), 1) if ir and ir.durata_media else "",
            round(float(ls.media), 1) if ls and ls.media else "",
            round(float(ls.minimo), 1) if ls and ls.minimo else "",
            round(float(ls.massimo), 1) if ls and ls.massimo else "",
        ])

    _style_sheet(ws, [20, 10, 24, 15, 18, 16, 14, 14])


@router.get("/letture")
async def export_letture(
    giorni: int = 30,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(require_admin),
):
    wb = openpyxl.Workbook()
    await _build_letture_sheet(wb.active, db, giorni)
    return _stream(wb, giorni)


@router.get("/irrigazioni")
async def export_irrigazioni(
    giorni: int = 30,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(require_admin),
):
    wb = openpyxl.Workbook()
    await _build_irrigazioni_sheet(wb.active, db, giorni)
    return _stream(wb, giorni)


@router.get("/completo")
async def export_completo(
    giorni: int = 30,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(require_admin),
):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()
    await _build_letture_sheet(ws1, db, giorni)
    await _build_irrigazioni_sheet(ws2, db, giorni)
    await _build_riepilogo_sheet(ws3, db, giorni)
    return _stream(wb, giorni)
